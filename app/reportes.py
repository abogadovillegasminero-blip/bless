import os
import sqlite3
from datetime import datetime
from io import BytesIO

from fastapi import APIRouter, Request
from fastapi.responses import RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates

from openpyxl import Workbook
from openpyxl.styles import Font

from app.auth import require_admin

router = APIRouter(prefix="/reportes", tags=["reportes"])
templates = Jinja2Templates(directory="templates")

DB_PATH = os.getenv("DB_PATH", "/tmp/bless.db")


def get_connection():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False, timeout=30)
    try:
        conn.execute("PRAGMA journal_mode=WAL;")
    except Exception:
        pass
    return conn


def _fetch_table_as_columns_and_rows(conn: sqlite3.Connection, table_name: str):
    cur = conn.cursor()

    cur.execute(f'PRAGMA table_info("{table_name}")')
    cols_info = cur.fetchall()
    columns = [c[1] for c in cols_info]  # nombre real de columnas

    if not columns:
        return [], []

    order_clause = ""
    if "id" in columns:
        order_clause = ' ORDER BY "id" ASC'

    select_cols = ", ".join([f'"{c}"' for c in columns])
    cur.execute(f'SELECT {select_cols} FROM "{table_name}"{order_clause}')
    rows = cur.fetchall()

    return columns, rows


def _autosize_worksheet(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)
        ws.column_dimensions[col_letter].width = max(10, min(max_len + 2, 60))


@router.get("/", response_class=None)
def ver_reportes(request: Request):
    user = require_admin(request)
    if isinstance(user, RedirectResponse):
        return user

    return templates.TemplateResponse(
        "reporte.html",
        {"request": request, "user": user},
    )


@router.get("/exportar-todo")
def exportar_todo(request: Request):
    user = require_admin(request)
    if isinstance(user, RedirectResponse):
        return user

    conn = get_connection()
    try:
        clientes_cols, clientes_rows = _fetch_table_as_columns_and_rows(conn, "clientes")
        pagos_cols, pagos_rows = _fetch_table_as_columns_and_rows(conn, "pagos")

        wb = Workbook()
        wb.remove(wb.active)  # quita hoja por defecto

        # Hoja CLIENTES
        ws_clientes = wb.create_sheet("CLIENTES")
        if clientes_cols:
            ws_clientes.append(clientes_cols)
            for cell in ws_clientes[1]:
                cell.font = Font(bold=True)
            for r in clientes_rows:
                ws_clientes.append([("" if v is None else v) for v in r])
            _autosize_worksheet(ws_clientes)
        else:
            ws_clientes.append(["Sin datos (tabla clientes no encontrada o sin columnas)"])
            ws_clientes["A1"].font = Font(bold=True)

        # Hoja PAGOS
        ws_pagos = wb.create_sheet("PAGOS")
        if pagos_cols:
            ws_pagos.append(pagos_cols)
            for cell in ws_pagos[1]:
                cell.font = Font(bold=True)
            for r in pagos_rows:
                ws_pagos.append([("" if v is None else v) for v in r])
            _autosize_worksheet(ws_pagos)
        else:
            ws_pagos.append(["Sin datos (tabla pagos no encontrada o sin columnas)"])
            ws_pagos["A1"].font = Font(bold=True)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"BLESS_export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    finally:
        try:
            conn.close()
        except Exception:
            pass
