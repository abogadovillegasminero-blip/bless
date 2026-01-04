# routers/reportes.py
from fastapi import APIRouter, Request, Depends
from fastapi.responses import StreamingResponse
from fastapi.templating import Jinja2Templates

from io import BytesIO
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font

# IMPORTA tu conexión a SQLite (ajusta si tu proyecto lo llama distinto)
# Debe existir porque ya tienes CRUD funcionando
from database import get_connection

# IMPORTA tu guard de admin (ajusta si tu proyecto lo llama distinto)
from routers.auth import require_admin

router = APIRouter(prefix="/reportes", tags=["reportes"])
templates = Jinja2Templates(directory="templates")


def _fetch_table_as_columns_and_rows(conn, table_name: str):
    """
    Lee columnas reales con PRAGMA table_info para no inventar nada,
    y trae todo el contenido ordenado por id (si existe).
    """
    cur = conn.cursor()

    # Columnas reales
    cur.execute(f'PRAGMA table_info("{table_name}")')
    cols_info = cur.fetchall()
    columns = [c[1] for c in cols_info]  # c[1] = name

    if not columns:
        return [], []

    # ORDER BY id si existe, si no, sin orden
    order_clause = ""
    if "id" in columns:
        order_clause = ' ORDER BY "id" ASC'

    # SELECT con comillas para evitar problemas con nombres
    select_cols = ", ".join([f'"{c}"' for c in columns])
    cur.execute(f'SELECT {select_cols} FROM "{table_name}"{order_clause}')
    rows = cur.fetchall()

    return columns, rows


def _autosize_worksheet(ws):
    """Auto-ajusta anchos de columna en base al contenido."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)
        # ancho razonable
        ws.column_dimensions[col_letter].width = max(10, min(max_len + 2, 60))


@router.get("")
def reportes_page(request: Request, user=Depends(require_admin)):
    return templates.TemplateResponse(
        "reportes.html",
        {"request": request, "user": user},
    )


@router.get("/exportar-todo")
def exportar_todo_excel(user=Depends(require_admin)):
    """
    Exporta TODO a Excel con 2 hojas: CLIENTES y PAGOS.
    Usa columnas reales desde SQLite (PRAGMA table_info).
    """
    conn = get_connection()

    try:
        clientes_cols, clientes_rows = _fetch_table_as_columns_and_rows(conn, "clientes")
        pagos_cols, pagos_rows = _fetch_table_as_columns_and_rows(conn, "pagos")

        wb = Workbook()

        # Elimina hoja por defecto
        default_ws = wb.active
        wb.remove(default_ws)

        # Hoja CLIENTES
        ws_clientes = wb.create_sheet("CLIENTES")
        if clientes_cols:
            ws_clientes.append(clientes_cols)
            for cell in ws_clientes[1]:
                cell.font = Font(bold=True)

            for r in clientes_rows:
                ws_clientes.append([("" if v is None else v) for v in r])

            _autosize_worksheet(ws_clientes)
        else:
            ws_clientes.append(["Sin datos (tabla clientes no encontrada o sin columnas)"])
            ws_clientes[1][0].font = Font(bold=True)

        # Hoja PAGOS
        ws_pagos = wb.create_sheet("PAGOS")
        if pagos_cols:
            ws_pagos.append(pagos_cols)
            for cell in ws_pagos[1]:
                cell.font = Font(bold=True)

            for r in pagos_rows:
                ws_pagos.append([("" if v is None else v) for v in r])

            _autosize_worksheet(ws_pagos)
        else:
            ws_pagos.append(["Sin datos (tabla pagos no encontrada o sin columnas)"])
            ws_pagos[1][0].font = Font(bold=True)

        # Generar archivo en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"BLESS_export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    finally:
        try:
            conn.close()
        except Exception:
            pass
